#!/usr/bin/env python3
"""
make_workbook.py - build the shareable spreadsheet from the committed CSVs.

Produces spreadsheet/Pittsburgh-Water-Deliberation-Record.xlsx in the same house style as the
Open the Books chart-data workbook: a Start-here tab, one tab per view, and a Sources tab. The
workbook is derived from 02-data/; edit the CSVs, then regenerate. Never hand-edit the .xlsx.

    python 03-harness/make_workbook.py
"""
import csv, json, pathlib, sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("make_workbook.py needs openpyxl (pip install openpyxl).")
    sys.exit(2)

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA = ROOT / "02-data"
OUT = ROOT / "spreadsheet" / "Pittsburgh-Water-Deliberation-Record.xlsx"
OUT.parent.mkdir(exist_ok=True)

INK = "1A1A1A"
ACCENT = "5B4B8A"
PAPER = "FBFAF7"
HEADFILL = PatternFill("solid", fgColor="EFEAF6")
TOTFILL = PatternFill("solid", fgColor="F2EFE8")
thin = Side(style="thin", color="D9D4C8")
BORDER = Border(bottom=thin)


def read(name):
    with open(DATA / name, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=ACCENT, size=10)
        cell.fill = HEADFILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_from_csv(wb, title, name, widths, total_label=None):
    rows = read(name)
    ws = wb.create_sheet(title)
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=_coerce(val))
            cell.alignment = Alignment(vertical="top", wrap_text=(c == len(row)))
    style_header(ws, 1, len(rows[0]))
    ws.freeze_panes = "A2"
    autosize(ws, widths)
    return ws, rows


def _coerce(v):
    if v is None or v == "" or v == "NA":
        return v
    try:
        return int(v)
    except ValueError:
        try:
            return float(v)
        except ValueError:
            return v


wb = Workbook()
wb.remove(wb.active)

# --- Start here -------------------------------------------------------------------
ws = wb.create_sheet("Start here")
lines = [
    ("The Pittsburgh Water Deliberation Record", 16, True, ACCENT),
    ("Chart data for two Public Source stories by Brian Nuckols", 11, False, INK),
    ("", 11, False, INK),
    ("Every resolution the Pittsburgh Water board voted, October 2023 to September 2025,", 11, False, INK),
    ("coded from the official minutes. These figures reproduce from the data with", 11, False, INK),
    ("    python 03-harness/build.py", 10, False, ACCENT),
    ("", 11, False, INK),
    ("THE HEADLINE FIGURES", 11, True, ACCENT),
]
r = 1
for text, size, bold, color in lines:
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = Font(size=size, bold=bold, color=color)
    r += 1
figures = [
    ("Board meetings analyzed", "24"),
    ("Resolutions voted", "228"),
    ("Unanimous votes", "223"),
    ("No votes / split votes", "0"),
    ("Abstentions (all Chair Sciulli, recusals)", "5"),
    ("Unanimity when all members voted", "100%"),
    ("Spending authorized", "$871.25 million"),
    ("Resolutions that drew a question", "72 of 228"),
    ("Resolutions that drew none", "156"),
    ("Public speakers, two years", "10 (across 5 of 24 meetings)"),
    ("Peer unanimity, random sample", "20 of 29 (69%)"),
    ("Questions at first post-publication meeting", "40 (about 7x the average)"),
]
for k, v in figures:
    ws.cell(row=r, column=1, value=k).font = Font(size=10)
    ws.cell(row=r, column=2, value=v).font = Font(size=10, bold=True)
    r += 1
r += 1
for note in [
    "TABS: Meeting-by-meeting - Resolutions questioned - Questions by member - Peer comparison - Sources",
    "The circulating draft figures 218 / 215 / $1.78B are wrong. The published figures are 228 / 223 / >$870M.",
    "This workbook is generated from 02-data/ by 03-harness/make_workbook.py. Edit the CSVs, then regenerate.",
]:
    ws.cell(row=r, column=1, value=note).font = Font(size=9, italic=True, color="4A4A4A")
    r += 1
autosize(ws, [46, 30])
ws.sheet_view.showGridLines = False

# --- data tabs --------------------------------------------------------------------
_, mrows = sheet_from_csv(wb, "Meeting-by-meeting", "meetings.csv",
                          [8, 13, 12, 10, 11, 10, 11, 12, 13, 42])
# append a totals row
ws_m = wb["Meeting-by-meeting"]
tr = len(mrows) + 1
def col(idx):  # sum an integer column across data rows
    return sum(int(mrows[i][idx]) for i in range(1, len(mrows)))
totals = ["TOTAL", "", col(2), col(3), col(4), col(5), col(6), col(7),
          round(sum(float(mrows[i][8]) for i in range(1, len(mrows))), 2), col(9), ""]
for c, v in enumerate(totals, 1):
    cell = ws_m.cell(row=tr, column=c, value=v)
    cell.font = Font(bold=True)
    cell.fill = TOTFILL

sheet_from_csv(wb, "Resolutions questioned", "questions_by_resolution.csv",
               [16, 8] + [10] * 8)
sheet_from_csv(wb, "Questions by member", "questions_by_member.csv",
               [22, 20, 15, 18, 15, 20])
sheet_from_csv(wb, "Peer comparison", "peer_comparison.csv",
               [30, 18, 15, 30, 8, 11, 11, 12, 12, 16, 16, 50])

# --- Sources ----------------------------------------------------------------------
cfg = json.loads((ROOT / "01-sources-archive" / "sources.json").read_text(encoding="utf-8"))
ws = wb.create_sheet("Sources")
ws.append(["id", "title", "url", "captured & hashed", "note"])
style_header(ws, 1, 5)
for s in cfg["sources"]:
    ws.append([s["id"], s["title"], s["url"],
               "yes" if s.get("mode") == "file" else "citation", s.get("note", "")])
ws.freeze_panes = "A2"
autosize(ws, [26, 44, 60, 16, 60])

wb.save(OUT)
print(f"wrote {OUT.relative_to(ROOT)} with {len(wb.sheetnames)} tabs: {', '.join(wb.sheetnames)}")
